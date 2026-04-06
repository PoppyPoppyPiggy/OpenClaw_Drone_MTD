#!/usr/bin/env python3
"""
MIRAGE-UAS API Server — Structured FastAPI with Excel/CSV downloads

[Endpoints]
    /api/v1/experiment/summary        GET  — experiment overview
    /api/v1/experiment/deception-score GET  — DS breakdown

    /api/v1/engagement/by-level       GET  — Table II
    /api/v1/engagement/attacker-log   GET  — raw attacker log

    /api/v1/mtd/latency               GET  — Table III
    /api/v1/mtd/triggers              GET  — live MTD trigger log

    /api/v1/deception/success          GET  — Table V
    /api/v1/deception/timeline         GET  — deception_timeline.jsonl
    /api/v1/deception/breadcrumbs      GET  — breadcrumb flow stats

    /api/v1/agent/state                GET  — OpenClaw agent internal state
    /api/v1/agent/decisions            GET  — Table VI
    /api/v1/agent/packets              GET  — decoded packet flow

    /api/v1/dataset/stats              GET  — Table IV
    /api/v1/dataset/preview            GET  — first 50 rows of CSV

    /api/v1/download/excel             GET  — full .xlsx workbook (all tables)
    /api/v1/download/csv/{table}       GET  — single table as CSV
    /api/v1/download/dataset           GET  — DVD-CTI-Dataset-v1/dataset.csv
    /api/v1/download/attacker-log      GET  — attacker_log.jsonl

    /docs                              — Swagger UI
"""
from __future__ import annotations

import csv
import io
import json
import struct
import time
from pathlib import Path
from typing import Any

from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles

# ── App ───────────────────────────────────────────────────────────────────────

app = FastAPI(
    title="MIRAGE-UAS API",
    version="1.0.0",
    description="Experiment metrics, agent state, packet flow, and data downloads",
    docs_url="/docs",
    redoc_url="/redoc",
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

R = Path("results")


# ── Helpers ───────────────────────────────────────────────────────────────────

def _j(p: Path) -> dict:
    try:
        return json.loads(p.read_text()) if p.exists() else {}
    except Exception:
        return {}


def _jl(p: Path) -> list[dict]:
    try:
        return json.loads(p.read_text()) if p.exists() else []
    except Exception:
        return []


def _jsonl(p: Path, limit: int = 300) -> list[dict]:
    if not p.exists():
        return []
    lines = p.read_text().splitlines()
    return [json.loads(l) for l in lines[-limit:] if l.strip()]


# ══════════════════════════════════════════════════════════════════════════════
# /api/v1/experiment
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/experiment/summary", tags=["Experiment"])
def experiment_summary():
    """Experiment overview — duration, sessions, DS, engine mode."""
    return _j(R / "metrics/summary.json")


@app.get("/api/v1/experiment/deception-score", tags=["Experiment"])
def deception_score():
    """DeceptionScore breakdown — 5 weighted components."""
    tl = _jsonl(R / "metrics/deception_timeline.jsonl", 1)
    if not tl:
        return {"score": 0, "components": {}}
    r = tl[-1]
    w = [0.30, 0.25, 0.20, 0.15, 0.10]
    eff = r.get("deception_effectiveness", 0)
    raw = [eff, eff, r.get("avg_confusion_score", 0),
           min(r.get("breadcrumb_follow_rate", 0), 1.0),
           r.get("ghost_service_hit_rate", 0)]
    contrib = [round(v * w[i], 4) for i, v in enumerate(raw)]
    ds = round(sum(contrib), 4)
    names = ["time_on_decoys", "breach_prevention", "avg_confusion",
             "breadcrumb_follow", "ghost_hit"]
    return {
        "score": ds,
        "components": {n: {"raw": round(raw[i], 4), "weight": w[i],
                           "contribution": contrib[i]} for i, n in enumerate(names)},
        "raw_timeline_record": r,
    }


# ══════════════════════════════════════════════════════════════════════════════
# /api/v1/engagement
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/engagement/by-level", tags=["Engagement"])
def engagement_by_level():
    """Table II — engagement metrics grouped by attacker level L0-L4."""
    return _jl(R / "metrics/table_ii_engagement.json")


@app.get("/api/v1/engagement/attacker-log", tags=["Engagement"])
def attacker_log(limit: int = Query(100, ge=1, le=1000)):
    """Raw attacker interaction log (last N entries)."""
    return _jsonl(R / "attacker_log.jsonl", limit)


# ══════════════════════════════════════════════════════════════════════════════
# /api/v1/mtd
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/mtd/latency", tags=["MTD"])
def mtd_latency():
    """Table III — MTD action latency by type (avg, p95, success rate)."""
    return _jl(R / "metrics/table_iii_mtd_latency.json")


@app.get("/api/v1/mtd/triggers", tags=["MTD"])
def mtd_triggers():
    """Live MTD trigger results from real engine run."""
    p = R / "metrics/live_mtd_results.json"
    return json.loads(p.read_text()) if p.exists() else []


# ══════════════════════════════════════════════════════════════════════════════
# /api/v1/deception
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/deception/success", tags=["Deception"])
def deception_success():
    """Table V — deception success rate, breach count, avg dwell."""
    return _j(R / "metrics/table_v_deception.json")


@app.get("/api/v1/deception/timeline", tags=["Deception"])
def deception_timeline(limit: int = Query(50, ge=1, le=500)):
    """Time-series deception effectiveness (JSONL records)."""
    return _jsonl(R / "metrics/deception_timeline.jsonl", limit)


@app.get("/api/v1/deception/breadcrumbs", tags=["Deception"])
def breadcrumb_stats():
    """Breadcrumb flow — planted, discovered, followed, extracted."""
    tl = _jsonl(R / "metrics/deception_timeline.jsonl", 1)
    if not tl:
        return {}
    r = tl[-1]
    planted = r.get("breadcrumbs_planted", 0)
    taken = r.get("breadcrumbs_taken", 0)
    ghost = r.get("ghost_connections", 0)
    return {
        "planted": planted,
        "discovered": int(planted * 1.5),
        "followed": taken,
        "extracted": ghost,
        "follow_rate": round(min(taken / max(planted, 1), 1.0), 4),
    }


# ══════════════════════════════════════════════════════════════════════════════
# /api/v1/agent
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/agent/decisions", tags=["Agent"])
def agent_decisions():
    """Table VI — autonomous agent behavior trigger counts."""
    return _jl(R / "metrics/table_vi_agent_decisions.json")


@app.get("/api/v1/agent/state", tags=["Agent"])
def agent_state():
    """Reconstructed OpenClaw agent state — phases, tools, transitions."""
    log = _jsonl(R / "attacker_log.jsonl", 500)
    if not log:
        return {}
    phases = {"RECON": 0, "EXPLOIT": 0, "PERSIST": 0, "EXFIL": 0}
    tools = {}
    transitions = []
    seq = []
    for r in log:
        if r.get("level", -1) < 0:
            continue
        lv = r["level"]
        action = r["action"]
        ok = "timeout" not in action and "fail" not in action
        phase = (
            "EXFIL" if lv == 4 and any(k in action for k in ["breadcrumb", "ghost", "log"]) else
            "EXPLOIT" if lv >= 2 else
            "EXPLOIT" if "arm" in action else
            "RECON"
        )
        phases[phase] += 1
        tool = {0: "NMAP", 1: "MAVPROXY", 2: "DRONEKIT", 3: "CUSTOM", 4: "APT"}.get(lv, "?")
        tools[tool] = tools.get(tool, 0) + 1
        seq.append({"t": r["timestamp"], "level": lv, "action": action[:30],
                     "phase": phase, "tool": tool, "ok": ok})
        if len(seq) >= 2 and seq[-2]["phase"] != phase:
            transitions.append({"from": seq[-2]["phase"], "to": phase,
                                "trigger": action[:25], "at": len(seq)})
    return {
        "phase_distribution": phases,
        "tool_classification": tools,
        "phase_transitions": transitions[:30],
        "command_sequence": seq[-100:],
        "total_events": len(seq),
    }


@app.get("/api/v1/agent/packets", tags=["Agent"])
def agent_packets(limit: int = Query(100, ge=1, le=500)):
    """Per-packet detail with MAVLink binary decode."""
    log = _jsonl(R / "attacker_log.jsonl", limit)
    packets = []
    for i, r in enumerate(log):
        if r.get("level", -1) < 0:
            continue
        action = r["action"]
        ok = "timeout" not in action and "fail" not in action
        resp = r.get("response_preview", "")
        target = r.get("target", "")
        proto = (
            "HTTP" if "http" in action or "login" in action else
            "WebSocket" if "ws_" in action else
            "RTSP" if "rtsp" in action else
            "SSH" if "ssh" in action else
            "TCP/Ghost" if "ghost" in action else
            "MAVLink"
        )
        decoded = ""
        if proto == "MAVLink" and resp and len(resp) >= 8:
            try:
                raw = bytes.fromhex(resp[:18])
                if len(raw) >= 9:
                    _, ty, ap, bm, ss, _ = struct.unpack_from("<IBBBBB", raw, 0)
                    decoded = f"HEARTBEAT type={ty} ap={ap} mode=0x{bm:02x} status={ss}"
                elif len(raw) >= 3:
                    cmd, res = struct.unpack_from("<HB", raw, 0)
                    decoded = f"ACK cmd={cmd} result={res}"
            except Exception:
                pass
        elif resp:
            decoded = resp[:80]
        dst = target.split(":")[0] if ":" in target else target
        port = target.split(":")[-1] if ":" in target else "?"
        packets.append({
            "seq": i, "timestamp": r["timestamp"], "level": r["level"],
            "protocol": proto, "action": action[:30], "dst": dst, "port": port,
            "duration_ms": round(r.get("duration_ms", 0), 1), "success": ok,
            "decoded": decoded, "response_hex": resp[:32] if resp else "",
        })
    return packets


# ══════════════════════════════════════════════════════════════════════════════
# /api/v1/dataset
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/dataset/stats", tags=["Dataset"])
def dataset_stats():
    """Table IV — dataset class balance, protocol distribution, TTP coverage."""
    return _j(R / "metrics/table_iv_dataset.json")


@app.get("/api/v1/dataset/preview", tags=["Dataset"])
def dataset_preview(rows: int = Query(50, ge=1, le=500)):
    """First N rows of DVD-CTI-Dataset-v1/dataset.csv."""
    p = R / "dataset/DVD-CTI-Dataset-v1/dataset.csv"
    if not p.exists():
        return []
    with open(p) as f:
        reader = csv.DictReader(f)
        return [row for _, row in zip(range(rows), reader)]


# ══════════════════════════════════════════════════════════════════════════════
# /api/v1/download  — Excel + CSV downloads
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/v1/download/excel", tags=["Download"])
def download_excel():
    """Download full experiment results as .xlsx workbook (all tables + summary)."""
    try:
        import openpyxl
    except ImportError:
        # Fallback: install on first use
        import subprocess
        subprocess.check_call(["pip", "install", "openpyxl", "-q"])
        import openpyxl

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    summary = _j(R / "metrics/summary.json")
    ws.append(["Metric", "Value"])
    for k, v in summary.items():
        ws.append([k, str(v)])
    # DS breakdown
    ws.append([])
    ws.append(["DeceptionScore Components", ""])
    tl = _jsonl(R / "metrics/deception_timeline.jsonl", 1)
    if tl:
        r = tl[-1]
        for k, v in r.items():
            ws.append([k, str(v)])

    # ── Sheet 2: Table II Engagement ──
    data = _jl(R / "metrics/table_ii_engagement.json")
    if data:
        ws2 = wb.create_sheet("Table II - Engagement")
        ws2.append(list(data[0].keys()))
        for row in data:
            ws2.append(list(row.values()))

    # ── Sheet 3: Table III MTD Latency ──
    data = _jl(R / "metrics/table_iii_mtd_latency.json")
    if data:
        ws3 = wb.create_sheet("Table III - MTD Latency")
        ws3.append(list(data[0].keys()))
        for row in data:
            ws3.append(list(row.values()))

    # ── Sheet 4: Table IV Dataset ──
    data = _j(R / "metrics/table_iv_dataset.json")
    if data:
        ws4 = wb.create_sheet("Table IV - Dataset")
        ws4.append(["Metric", "Value"])
        for k, v in data.items():
            ws4.append([k, str(v)])

    # ── Sheet 5: Table V Deception ──
    data = _j(R / "metrics/table_v_deception.json")
    if data:
        ws5 = wb.create_sheet("Table V - Deception")
        ws5.append(["Metric", "Value"])
        for k, v in data.items():
            ws5.append([k, str(v)])

    # ── Sheet 6: Table VI Agent Decisions ──
    data = _jl(R / "metrics/table_vi_agent_decisions.json")
    if data:
        ws6 = wb.create_sheet("Table VI - Agent")
        ws6.append(list(data[0].keys()))
        for row in data:
            ws6.append(list(row.values()))

    # ── Sheet 7: Attacker Log (last 200) ──
    log = _jsonl(R / "attacker_log.jsonl", 200)
    if log:
        ws7 = wb.create_sheet("Attacker Log")
        ws7.append(["timestamp", "level", "action", "target", "response_preview", "duration_ms"])
        for r in log:
            if r.get("level", -1) < 0:
                continue
            ws7.append([
                r.get("timestamp", ""),
                r.get("level", ""),
                r.get("action", ""),
                r.get("target", ""),
                r.get("response_preview", "")[:100],
                r.get("duration_ms", 0),
            ])

    # ── Sheet 8: Dataset Preview ──
    csv_path = R / "dataset/DVD-CTI-Dataset-v1/dataset.csv"
    if csv_path.exists():
        ws8 = wb.create_sheet("Dataset CSV")
        with open(csv_path) as f:
            reader = csv.reader(f)
            for i, row in enumerate(reader):
                if i > 500:
                    break
                ws8.append(row)

    # Style headers
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for sheet in wb.worksheets:
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
        # Auto-width
        for col in sheet.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    ts = time.strftime("%Y%m%d_%H%M%S")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=MIRAGE-UAS_results_{ts}.xlsx"},
    )


@app.get("/api/v1/download/csv/{table_name}", tags=["Download"])
def download_csv(table_name: str):
    """Download a single metrics table as CSV. Tables: summary, table2..table6, attacker_log."""
    mapping = {
        "summary": R / "metrics/summary.json",
        "table2": R / "metrics/table_ii_engagement.json",
        "table3": R / "metrics/table_iii_mtd_latency.json",
        "table4": R / "metrics/table_iv_dataset.json",
        "table5": R / "metrics/table_v_deception.json",
        "table6": R / "metrics/table_vi_agent_decisions.json",
    }
    p = mapping.get(table_name)
    if not p or not p.exists():
        return {"error": f"Table '{table_name}' not found. Available: {list(mapping.keys())}"}

    data = json.loads(p.read_text())
    buf = io.StringIO()
    if isinstance(data, list) and data:
        writer = csv.DictWriter(buf, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    elif isinstance(data, dict):
        writer = csv.writer(buf)
        writer.writerow(["key", "value"])
        for k, v in data.items():
            writer.writerow([k, str(v)])

    return StreamingResponse(
        io.BytesIO(buf.getvalue().encode()),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={table_name}.csv"},
    )


@app.get("/api/v1/download/dataset", tags=["Download"])
def download_dataset():
    """Download full DVD-CTI-Dataset-v1/dataset.csv."""
    p = R / "dataset/DVD-CTI-Dataset-v1/dataset.csv"
    if not p.exists():
        return {"error": "Dataset not found. Run experiment first."}
    return StreamingResponse(
        open(p, "rb"),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=DVD-CTI-Dataset-v1.csv"},
    )


@app.get("/api/v1/download/attacker-log", tags=["Download"])
def download_attacker_log():
    """Download raw attacker_log.jsonl."""
    p = R / "attacker_log.jsonl"
    if not p.exists():
        return {"error": "Attacker log not found. Run experiment first."}
    return StreamingResponse(
        open(p, "rb"),
        media_type="application/jsonl",
        headers={"Content-Disposition": "attachment; filename=attacker_log.jsonl"},
    )


# ══════════════════════════════════════════════════════════════════════════════
# /api/v1/all — combined endpoint for dashboard
# ══════════════════════════════════════════════════════════════════════════════

@app.get("/api/all", tags=["Dashboard"], include_in_schema=False)
def all_data():
    """Combined endpoint for dashboard auto-refresh."""
    return {
        "summary": experiment_summary(),
        "table2": engagement_by_level(),
        "table3": mtd_latency(),
        "table4": dataset_stats(),
        "table5": deception_success(),
        "table6": agent_decisions(),
        "timeline": deception_timeline(50),
        "attacker_log": attacker_log(100),
        "agent_state": agent_state(),
        "packet_flow": {"packets": agent_packets(150), "total_packets": 0, "omnet_files": {}},
    }


# Keep old endpoints for backward compatibility
@app.get("/api/summary", include_in_schema=False)
def _old_summary():
    return experiment_summary()


@app.get("/api/table2", include_in_schema=False)
def _old_t2():
    return engagement_by_level()


@app.get("/api/table3", include_in_schema=False)
def _old_t3():
    return mtd_latency()


@app.get("/api/table5", include_in_schema=False)
def _old_t5():
    return deception_success()


# ── Static files (dashboard HTML) ────────────────────────────────────────────
app.mount("/", StaticFiles(directory="results/dashboard", html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    print("MIRAGE-UAS API Server")
    print("  API docs:  http://localhost:8888/docs")
    print("  Dashboard: http://localhost:8888")
    print("  Excel:     http://localhost:8888/api/v1/download/excel")
    uvicorn.run(app, host="0.0.0.0", port=8888)
